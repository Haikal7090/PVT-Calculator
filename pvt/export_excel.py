from __future__ import annotations

from pathlib import Path
from typing import Iterable, List
from dataclasses import asdict

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import FieldInfo, MethodSelection, PVTInput, PVTPoint
from .verification import physical_consistency_checks, run_textbook_verification

DARK = "1F4E78"
BLUE = "5B9BD5"
LIGHT_BLUE = "D9EAF7"
GREEN = "70AD47"
LIGHT_GREEN = "E2F0D9"
GOLD = "FFC000"
LIGHT_GOLD = "FFF2CC"
GRAY = "E7E6E6"
RED = "C00000"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="B7B7B7")


def _title(ws, cell_range: str, text: str, fill: str = DARK) -> None:
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = text
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(color=WHITE, bold=True, size=14)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _section_header(ws, row: int, start_col: int, end_col: int, text: str, fill: str = BLUE) -> None:
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    cell = ws.cell(row, start_col, text)
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(color=WHITE, bold=True)
    cell.alignment = Alignment(horizontal="center")


def _border_region(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _autosize(ws, minimum: int = 10, maximum: int = 28) -> None:
    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_letter].width = min(max(length + 2, minimum), maximum)


def export_workbook(
    destination: str | Path,
    field_info: FieldInfo,
    data: PVTInput,
    methods: MethodSelection,
    point: PVTPoint,
    table: List[PVTPoint],
) -> Path:
    destination = Path(destination)
    wb = Workbook()
    interface = wb.active
    interface.title = "Interface"
    calculator = wb.create_sheet("PVT Calculator")
    table_ws = wb.create_sheet("PVT Table")
    verification = wb.create_sheet("Textbook Verification")
    consistency = wb.create_sheet("Physical Consistency")
    charts = wb.create_sheet("Charts")
    methods_ws = wb.create_sheet("Methods")

    # ------------------------------------------------------------------
    # Interface
    # ------------------------------------------------------------------
    _title(interface, "A1:H2", "PVT CALCULATOR — OIL, GAS & BRINE PROPERTIES", GREEN)
    _section_header(interface, 4, 1, 4, "Field Information", GREEN)
    field_rows = [
        ("Field Name", field_info.field_name),
        ("Company", field_info.company),
        ("Location", field_info.location),
        ("Engineer", field_info.engineer),
    ]
    for idx, (label, value) in enumerate(field_rows, start=5):
        interface.cell(idx, 1, label)
        interface.cell(idx, 2, value)

    _section_header(interface, 10, 1, 4, "General Reservoir Data")
    general_rows = [
        ("Reservoir Temperature", data.reservoir_temperature_f, "°F"),
        ("Initial Reservoir Pressure", data.initial_reservoir_pressure_psia, "psia"),
        ("Evaluation Pressure", data.evaluation_pressure_psia, "psia"),
        ("Standard Pressure", data.standard_pressure_psia, "psia"),
        ("Gas Specific Gravity", data.gas_specific_gravity, "air = 1"),
    ]
    for idx, row in enumerate(general_rows, start=11):
        interface.cell(idx, 1, row[0]); interface.cell(idx, 2, row[1]); interface.cell(idx, 3, row[2])

    _section_header(interface, 17, 1, 4, "Oil Data")
    oil_rows = [
        ("Oil API", data.oil_api, "°API"),
        ("Bubble Point Pressure", data.bubble_point_pressure_psia, "psia"),
        ("Separator Pressure", data.separator_pressure_psia, "psia"),
        ("Separator Temperature", data.separator_temperature_f, "°F"),
    ]
    for idx, row in enumerate(oil_rows, start=18):
        interface.cell(idx, 1, row[0]); interface.cell(idx, 2, row[1]); interface.cell(idx, 3, row[2])

    _section_header(interface, 23, 1, 4, "Impurities and Brine")
    impurity_rows = [
        ("CO₂", data.co2_mol_pct, "mol%"),
        ("H₂S", data.h2s_mol_pct, "mol%"),
        ("N₂", data.n2_mol_pct, "mol%"),
        ("TDS", data.tds_weight_pct, "wt%"),
        ("Brine Condition", data.brine_condition, ""),
    ]
    for idx, row in enumerate(impurity_rows, start=24):
        interface.cell(idx, 1, row[0]); interface.cell(idx, 2, row[1]); interface.cell(idx, 3, row[2])

    _section_header(interface, 4, 5, 8, "Selected Correlations", GOLD)
    for idx, (name, value) in enumerate(asdict(methods).items(), start=5):
        interface.cell(idx, 5, name.replace("_", " ").title())
        interface.cell(idx, 6, value.replace("_", " ").title())
    interface.sheet_view.showGridLines = False
    _border_region(interface, 4, 28, 1, 8)

    # ------------------------------------------------------------------
    # Calculator
    # ------------------------------------------------------------------
    _title(calculator, "A1:H2", "PVT CALCULATOR RESULTS", DARK)
    _section_header(calculator, 4, 1, 4, f"Properties at {point.pressure_psia:.2f} psia", GOLD)
    result_rows = [
        ("Condition", point.condition, ""),
        ("Bo", point.bo_rb_stb, "RB/STB"),
        ("Rs", point.rs_scf_stb / 1000.0, "Mscf/STB"),
        ("Bg", point.bg_rb_mscf, "RB/Mscf"),
        ("Eg", point.eg_mscf_rb, "Mscf/RB"),
        ("Bw", point.bw_rb_stb, "RB/STB"),
        ("Rsw", point.rsw_scf_stb, "scf/STB"),
        ("H₂O in Gas", point.water_content_lbm_mmscf, "lbm/MMscf"),
        ("Z-factor", point.z_factor, "vol/vol"),
        ("Oil Density", point.oil_density_lbm_ft3, "lbm/ft³"),
        ("Gas Density", point.gas_density_lbm_ft3, "lbm/ft³"),
        ("Brine Density", point.brine_density_lbm_ft3, "lbm/ft³"),
        ("Oil Viscosity", point.oil_viscosity_cp, "cp"),
        ("Dead Oil Viscosity", point.dead_oil_viscosity_cp, "cp"),
        ("Gas Viscosity", point.gas_viscosity_cp, "cp"),
        ("Brine Viscosity", point.brine_viscosity_cp, "cp"),
        ("Co × 10⁵", point.co_1_psi * 1e5, "1/psi"),
        ("Cg × 10⁵", point.cg_1_psi * 1e5, "1/psi"),
        ("Cw × 10⁵", point.cw_1_psi * 1e5, "1/psi"),
    ]
    for idx, (label, value, unit) in enumerate(result_rows, start=5):
        calculator.cell(idx, 1, label)
        calculator.cell(idx, 2, value)
        calculator.cell(idx, 3, unit)

    _section_header(calculator, 4, 5, 8, "Critical Condition", BLUE)
    critical_rows = [
        ("Tpc", point.critical.tpc_r, "°R"),
        ("Ppc", point.critical.ppc_psia, "psia"),
        ("Corrected Tpc", point.critical.corrected_tpc_r, "°R"),
        ("Corrected Ppc", point.critical.corrected_ppc_psia, "psia"),
        ("Tpr", point.critical.tpr, "dimensionless"),
        ("Ppr", point.critical.ppr, "dimensionless"),
    ]
    for idx, (label, value, unit) in enumerate(critical_rows, start=5):
        calculator.cell(idx, 5, label)
        calculator.cell(idx, 6, value)
        calculator.cell(idx, 7, unit)
    _border_region(calculator, 4, 23, 1, 8)
    calculator.sheet_view.showGridLines = False

    # ------------------------------------------------------------------
    # Table
    # ------------------------------------------------------------------
    headers = [
        "Pressure (psia)", "Condition", "Bo (RB/STB)", "Rs (Mscf/STB)", "Bg (RB/Mscf)",
        "Eg (Mscf/RB)", "Bw (RB/STB)", "Rsw (scf/STB)", "H2O Gas (lbm/MMscf)",
        "Z", "Oil Density", "Gas Density", "Brine Density", "Oil Viscosity", "Gas Viscosity",
        "Brine Viscosity", "Co x1e5", "Cg x1e5", "Cw x1e5",
    ]
    table_ws.append(headers)
    for item in table:
        table_ws.append([
            item.pressure_psia, item.condition, item.bo_rb_stb, item.rs_scf_stb / 1000.0,
            item.bg_rb_mscf, item.eg_mscf_rb, item.bw_rb_stb, item.rsw_scf_stb,
            item.water_content_lbm_mmscf, item.z_factor, item.oil_density_lbm_ft3,
            item.gas_density_lbm_ft3, item.brine_density_lbm_ft3, item.oil_viscosity_cp,
            item.gas_viscosity_cp, item.brine_viscosity_cp, item.co_1_psi * 1e5,
            item.cg_1_psi * 1e5, item.cw_1_psi * 1e5,
        ])
    for cell in table_ws[1]:
        cell.fill = PatternFill("solid", fgColor=DARK)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    table_ws.freeze_panes = "A2"
    table_ws.auto_filter.ref = table_ws.dimensions
    _border_region(table_ws, 1, table_ws.max_row, 1, table_ws.max_column)

    # ------------------------------------------------------------------
    # Textbook equation verification
    # ------------------------------------------------------------------
    _title(verification, "A1:N2", "TEXTBOOK EQUATION VERIFICATION", GREEN)
    verification["A3"] = (
        "Purpose: verify that the coded equations reproduce published textbook examples. "
        "This is code verification, not experimental validation against a field-fluid PVT laboratory report."
    )
    verification.merge_cells("A3:N3")
    verification["A3"].alignment = Alignment(wrap_text=True, vertical="center")
    verification_headers = [
        "Test ID", "Category", "Function", "Calculated", "Expected", "Unit",
        "Abs. Error", "Error (%)", "Tolerance (%)", "Status", "Textbook",
        "Location", "Equation / Example", "Input and Notes",
    ]
    verification.append(verification_headers)
    textbook_rows = run_textbook_verification()
    for row in textbook_rows:
        combined_note = row["input_summary"]
        if row.get("note"):
            combined_note += " | " + str(row["note"])
        verification.append([
            row["test_id"], row["category"], row["function_name"], row["calculated"],
            row["expected"], row["unit"], row["absolute_error"], row["percent_error"],
            row["tolerance_pct"], row["status"], row["source"], row["source_location"],
            row["equation_or_example"], combined_note,
        ])
    for cell in verification[4]:
        cell.fill = PatternFill("solid", fgColor=DARK)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_idx in range(5, verification.max_row + 1):
        status_cell = verification.cell(row_idx, 10)
        passed = status_cell.value == "PASS"
        status_cell.fill = PatternFill("solid", fgColor=LIGHT_GREEN if passed else "F4CCCC")
        status_cell.font = Font(bold=True, color="006100" if passed else RED)
    _border_region(verification, 4, verification.max_row, 1, 14)
    verification.freeze_panes = "A5"

    # ------------------------------------------------------------------
    # Physical consistency checks for the user's current PVT case
    # ------------------------------------------------------------------
    _title(consistency, "A1:F2", "PHYSICAL CONSISTENCY CHECKS", BLUE)
    consistency["A3"] = (
        "These checks assess internal identities and expected black-oil behavior. "
        "They do not replace comparison with laboratory PVT data."
    )
    consistency.merge_cells("A3:F3")
    consistency["A3"].alignment = Alignment(wrap_text=True, vertical="center")
    consistency.append(["Check ID", "Check", "Expected Behavior", "Observed", "Status", "Note"])
    physical_rows = physical_consistency_checks(data, methods, point, table)
    for row in physical_rows:
        consistency.append([
            row["check_id"], row["check"], row["expected"], row["observed"],
            row["status"], row["note"],
        ])
    for cell in consistency[4]:
        cell.fill = PatternFill("solid", fgColor=DARK)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_idx in range(5, consistency.max_row + 1):
        status_cell = consistency.cell(row_idx, 5)
        passed = status_cell.value == "PASS"
        status_cell.fill = PatternFill("solid", fgColor=LIGHT_GREEN if passed else "F4CCCC")
        status_cell.font = Font(bold=True, color="006100" if passed else RED)
    _border_region(consistency, 4, consistency.max_row, 1, 6)
    consistency.freeze_panes = "A5"

    # ------------------------------------------------------------------
    # Charts
    # ------------------------------------------------------------------
    _title(charts, "A1:P2", "PVT CHART BANK", DARK)
    # Every calculated PVT property is exported as an individual chart.
    # Tuple format: (starting row, chart title, source column, y-axis title).
    chart_specs = [
        (3, "Bo vs Pressure", 3, "Bo (RB/STB)"),
        (3, "Rs vs Pressure", 4, "Rs (Mscf/STB)"),
        (21, "Bg vs Pressure", 5, "Bg (RB/Mscf)"),
        (21, "Eg vs Pressure", 6, "Eg (Mscf/RB)"),
        (39, "Bw vs Pressure", 7, "Bw (RB/STB)"),
        (39, "Rsw vs Pressure", 8, "Rsw (scf/STB)"),
        (57, "Water Content in Gas vs Pressure", 9, "H2O (lbm/MMscf)"),
        (57, "Z-factor vs Pressure", 10, "Z-factor"),
        (75, "Oil Density vs Pressure", 11, "Oil Density (lbm/ft³)"),
        (75, "Gas Density vs Pressure", 12, "Gas Density (lbm/ft³)"),
        (93, "Brine Density vs Pressure", 13, "Brine Density (lbm/ft³)"),
        (93, "Oil Viscosity vs Pressure", 14, "Oil Viscosity (cp)"),
        (111, "Gas Viscosity vs Pressure", 15, "Gas Viscosity (cp)"),
        (111, "Brine Viscosity vs Pressure", 16, "Brine Viscosity (cp)"),
        (129, "Oil Compressibility vs Pressure", 17, "Co × 10⁵ (1/psi)"),
        (129, "Gas Compressibility vs Pressure", 18, "Cg × 10⁵ (1/psi)"),
        (147, "Water Compressibility vs Pressure", 19, "Cw × 10⁵ (1/psi)"),
    ]
    data_max_row = table_ws.max_row
    for index, (start_row, title, column, y_title) in enumerate(chart_specs):
        chart = LineChart()
        chart.title = title
        chart.style = 13
        chart.y_axis.title = y_title
        chart.x_axis.title = "Pressure (psia)"
        chart.height = 7.2
        chart.width = 13.5
        chart.legend = None
        chart.add_data(
            Reference(table_ws, min_col=column, min_row=1, max_row=data_max_row),
            titles_from_data=True,
        )
        chart.set_categories(
            Reference(table_ws, min_col=1, min_row=2, max_row=data_max_row)
        )
        # Two-column chart layout: left at A, right at J.
        anchor_column = "A" if index % 2 == 0 else "J"
        charts.add_chart(chart, f"{anchor_column}{start_row}")
    charts.sheet_view.showGridLines = False

    # ------------------------------------------------------------------
    # Methods
    # ------------------------------------------------------------------
    methods_ws.append(["Category", "Selected Method", "Notes"])
    for name, value in asdict(methods).items():
        methods_ws.append([name, value, "All inputs use field units; mole impurities are converted to fractions internally."])
    methods_ws.sheet_state = "hidden"

    for ws in (interface, calculator, table_ws, verification, consistency, charts):
        _autosize(ws)
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = "0.000000"
                cell.alignment = Alignment(vertical="center", wrap_text=True)

    wb.save(destination)
    return destination
